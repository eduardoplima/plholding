"""Importa os dados reais da planilha docs/Gestão.xlsx para o banco do sistema.

Uso (a partir de backend/, com o Postgres no ar e o .env carregado):

    set -a && . ./.env && set +a
    PYTHONPATH=. uv run python scripts/import_gestao.py --reset

Decisões de importação (acordadas com o usuário):
  - Linhas bagunçadas: limpa datas/durações/nomes via regex; pula o que não
    tem data válida e lista no relatório final.
  - Imóveis alugados inteiros (aba Aluguéis): cria Property(whole) + Unidade
    "Inteiro" com o valor como base_rent; NÃO cria contrato.
  - Dívidas: importa as 4 nomeadas (saldo devedor = valor contratado); pula as
    linhas sem nome.
  - --reset limpa as tabelas de domínio (mantém users) antes de importar.
"""
from __future__ import annotations

import argparse
import os
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

os.environ.setdefault('SKIP_SEED', '1')  # não rodar seed/create_all ao importar

import openpyxl
from sqlalchemy import delete, func, select

from app.main import (  # noqa: E402
    Debt, DebtKind, Document, Expense, Lease, LeaseStatus, LeaseTenant,
    Property, PropertyKind, RentCharge, RentalMode, SessionLocal, Tenant, Unit,
    add_months, recompute_lease, recompute_units,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_FILE = REPO_ROOT / 'docs' / 'Gestão.xlsx'

# Mapa canônico dos imóveis: nome -> (kind, rental_mode). Aliases resolvidos em norm_property().
PROPERTY_KIND = {
    'Condomínio 1': (PropertyKind.condominio, RentalMode.by_unit),
    'Condomínio 2': (PropertyKind.condominio, RentalMode.by_unit),
    'Condomínio 3': (PropertyKind.condominio, RentalMode.by_unit),
    'Condomínio 4': (PropertyKind.condominio, RentalMode.whole),
    'Casa Tabatinga': (PropertyKind.casa, RentalMode.whole),
    'Fazenda Poço Branco': (PropertyKind.fazenda, RentalMode.whole),
    'Fazenda Mossoró': (PropertyKind.fazenda, RentalMode.whole),
    'Clínica Mossoró': (PropertyKind.clinica, RentalMode.whole),
    'Restaurante': (PropertyKind.comercial, RentalMode.whole),
}
PROPERTY_ALIAS = {'cond 4': 'Condomínio 4', 'clínica': 'Clínica Mossoró', 'clinica': 'Clínica Mossoró'}

DEBT_PROPERTY = {
    'Consignado Reforma Condomínio 3': 'Condomínio 3',
    'Consignado Compra Casa Tabatinga': 'Casa Tabatinga',
    'Financiamento Cond 4': 'Condomínio 4',
    'Poço branco': 'Fazenda Poço Branco',
}


def dec(v) -> Decimal | None:
    if v is None or str(v).strip() == '':
        return None
    return Decimal(str(v)).quantize(Decimal('0.01'))


def norm_property(name: str) -> str:
    n = str(name).strip()
    return PROPERTY_ALIAS.get(n.lower(), n)


def clean_name(name: str) -> str:
    """Remove dígitos soltos no fim do nome (lixo da planilha)."""
    return re.sub(r'\s*\d+\s*$', '', str(name).strip()).strip()


def parse_date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    m = re.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', str(v or ''))
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    try:
        return date(y, mo, d)
    except ValueError:
        return None


def parse_months(v) -> int | None:
    s = str(v or '').lower()
    m = re.search(r'(\d+)', s)
    if not m:
        return None
    n = int(m.group(1))
    if 'ano' in s:
        return n * 12
    if 'mes' in s or 'mês' in s:
        return n
    return None


def reset_domain(db):
    for model in (Document, RentCharge, LeaseTenant, Lease, Expense, Unit, Debt, Property, Tenant):
        db.execute(delete(model))
    db.commit()


def run(path: Path, do_reset: bool):
    wb = openpyxl.load_workbook(path, data_only=True)
    db = SessionLocal()
    report: list[str] = []
    try:
        if do_reset:
            reset_domain(db)
            report.append('Reset: tabelas de domínio limpas (users preservados).')

        # ---- 1) Properties (Patrimônio + Aluguéis) ----
        props: dict[str, Property] = {}
        rents_whole: dict[str, Decimal] = {}

        for row in wb['Alugueis'].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = norm_property(row[0])
            if dec(row[1]) is not None:
                rents_whole[name] = dec(row[1])

        for row in wb['Patrimônio'].iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            name = norm_property(row[0])
            kind, mode = PROPERTY_KIND.get(name, (PropertyKind.outro, RentalMode.whole))
            props[name] = Property(name=name, kind=kind, rental_mode=mode, market_value=dec(row[1]))

        # imóveis que só aparecem em Aluguéis (ex.: Restaurante)
        for name in rents_whole:
            if name not in props:
                kind, mode = PROPERTY_KIND.get(name, (PropertyKind.outro, RentalMode.whole))
                props[name] = Property(name=name, kind=kind, rental_mode=mode, market_value=None)

        for p in props.values():
            db.add(p)
        db.flush()

        # unidade "Inteiro" para imóveis whole
        for name, p in props.items():
            if p.rental_mode == RentalMode.whole:
                db.add(Unit(property_id=p.id, name='Inteiro', base_rent=rents_whole.get(name, Decimal('0.00'))))
        db.flush()

        # ---- 2) Condomínios -> Units + Tenants + Leases ----
        units: dict[tuple[str, str], Unit] = {}
        tenants: dict[str, Tenant] = {}
        seen_leases: set[tuple[int, str, str]] = set()
        leases_created = 0

        def get_tenant(name: str) -> Tenant:
            key = name.lower()
            if key not in tenants:
                t = Tenant(full_name=name)
                db.add(t)
                db.flush()
                tenants[key] = t
            return tenants[key]

        for i, row in enumerate(wb['Condomínios'].iter_rows(min_row=2, values_only=True), start=2):
            if not row or row[0] is None:
                continue
            cond_raw, unit_raw, valor, data_raw, dur_raw, loc_raw = (list(row) + [None] * 6)[:6]
            try:
                cond_name = f'Condomínio {int(float(cond_raw))}'
            except (TypeError, ValueError):
                report.append(f'  linha {i}: condomínio inválido ({cond_raw!r}) — pulada')
                continue
            prop = props.get(cond_name)
            if prop is None or prop.rental_mode != RentalMode.by_unit:
                report.append(f'  linha {i}: {cond_name} não é condomínio by_unit — pulada')
                continue

            start = parse_date(data_raw)
            months = parse_months(dur_raw)
            unit_name = str(unit_raw).strip() if unit_raw is not None else ''
            bad_unit = unit_name == '' or re.fullmatch(r'\d+(\.0)?', unit_name) is not None
            if start is None or months is None or not loc_raw or bad_unit:
                report.append(f'  linha {i}: dados incompletos (unidade={unit_raw!r}, data={data_raw!r}, dur={dur_raw!r}, loc={loc_raw!r}) — pulada')
                continue

            ukey = (cond_name, unit_name)
            if ukey not in units:
                u = Unit(property_id=prop.id, name=unit_name, base_rent=dec(valor) or Decimal('0.00'))
                db.add(u)
                db.flush()
                units[ukey] = u
            unit = units[ukey]

            names = [clean_name(n) for n in str(loc_raw).split(',')]
            names = [n for n in names if n]
            if not names:
                report.append(f'  linha {i}: sem locatário válido — pulada')
                continue

            lkey = (unit.id, start.isoformat(), names[0].lower())
            if lkey in seen_leases:
                continue
            seen_leases.add(lkey)

            end = add_months(start, months) - timedelta(days=1)
            lease = Lease(
                unit_id=unit.id, start_date=start, duration_months=months, end_date=end,
                monthly_rent=dec(valor) or Decimal('0.00'), due_day=min(start.day, 28),
            )
            recompute_lease(lease)
            db.add(lease)
            db.flush()
            for j, nm in enumerate(names):
                db.add(LeaseTenant(lease_id=lease.id, tenant_id=get_tenant(nm).id, is_primary=(j == 0)))
            leases_created += 1

        # ---- 3) Dívidas ----
        debts_created = 0
        for i, row in enumerate(wb['Dívidas'].iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0] or not str(row[0]).strip():
                if row and (row[1] or (len(row) > 2 and row[2])):
                    report.append(f'  dívida linha {i}: sem nome ({list(row)!r}) — pulada')
                continue
            name = str(row[0]).strip()
            principal = dec(row[1]) or Decimal('0.00')
            parcela = dec(row[2]) if len(row) > 2 else None
            prop_name = DEBT_PROPERTY.get(name)
            kind = DebtKind.financiamento if 'financ' in name.lower() else DebtKind.consignado if 'consig' in name.lower() else DebtKind.outro
            db.add(Debt(
                name=name, kind=kind, principal_amount=principal, outstanding_balance=principal,
                installment_amount=parcela, property_id=props[prop_name].id if prop_name in props else None,
            ))
            debts_created += 1

        recompute_units(db)
        db.commit()

        report.append('')
        report.append('=== RESUMO ===')
        report.append(f'Imóveis:     {len(props)}')
        report.append(f'Unidades:    {db.scalar(select(func.count(Unit.id)))}')
        report.append(f'Inquilinos:  {len(tenants)}')
        report.append(f'Contratos:   {leases_created}')
        report.append(f'Dívidas:     {debts_created}')
        report.append('Despesas:    0 (aba Despesas sem valores na planilha)')
    finally:
        db.close()

    print('\n'.join(report))


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--file', default=str(DEFAULT_FILE))
    ap.add_argument('--reset', action='store_true', help='limpa tabelas de domínio antes de importar')
    args = ap.parse_args()
    run(Path(args.file), args.reset)
