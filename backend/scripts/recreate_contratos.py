"""Recria unidades canônicas, contratos (leases) dos inquilinos atuais e gera as
cobranças mensais — pré-requisito da conciliação bancária (precisa do "esperado").

Fontes de valor por unidade:
  - Cond 1/2: controle_alugueis.xlsx
  - Cond 3: Alugueis unidades.xlsx
  - Imóveis inteiros: aba Aluguéis da Gestão.xlsx (Cond 4=4000, Clínica=10000, Restaurante=3000)

Inquilino atual por unidade = mesmo critério de import_inquilinos.py (controle +
overrides dos contratos 2026). Liga aos Tenants já cadastrados (por nome).

Uso (de backend/, Postgres no ar e .env carregado):
    set -a && . ./.env && set +a
    PYTHONPATH=. uv run python scripts/recreate_contratos.py
"""
from __future__ import annotations

import os
import re
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

os.environ.setdefault('SKIP_SEED', '1')

import openpyxl
from sqlalchemy import delete, func, select

from app.main import (  # noqa: E402
    ChargeStatus, Lease, LeaseStatus, LeaseTenant, Property, RentCharge, RentalMode,
    SessionLocal, Tenant, Unit, add_months, month_start, recompute_charge,
    recompute_lease, recompute_units,
)

REPO = Path(__file__).resolve().parents[2]
CONTROLE = REPO / 'docs' / 'Aluguel' / 'Antigos' / 'Condominio_2' / 'controle_alugueis.xlsx'
COND3_VALS = REPO / 'docs' / 'Aluguel' / 'Antigos' / 'Condominio_3' / 'Alugueis unidades.xlsx'

UP_TO = date(2026, 5, 1)  # gera cobranças até maio/2026 (cobre o extrato)

# inquilino atual por unidade vindo dos contratos 2026 (prioridade) — nome + data início
CONTRATOS_2026 = {
    (1, 1): (['Isaque Italo Medeiros Dantas'], date(2026, 3, 10)),
    (1, 6): (['Enrique Bezerra de Souza'], date(2026, 4, 13)),
    (2, 3): (['João Vitor de Oliveira Alves', 'Alice'], date(2026, 3, 10)),
    (3, 1): (['Eduarda Fabricio Passo'], date(2026, 2, 26)),
    (3, 2): (['Thiago Duarte Coelho de Souza Dantas'], date(2026, 5, 5)),
    (3, 8): (['Erik Leandro Lucas de Oliveira'], date(2026, 3, 15)),
}
WHOLE = {  # nome do imóvel -> (inquilino, valor)
    'Condomínio 4': ('Ramon e Scheila Pinheiro', Decimal('4000.00')),
    'Restaurante': ('Flávio', Decimal('3000.00')),
    'Clínica Mossoró': (None, Decimal('10000.00')),
}


def norm(s: str) -> str:
    return re.sub(r'\s+', ' ', str(s)).strip()


def read_cond12() -> dict[tuple[int, int], tuple[list[str], Decimal]]:
    wb = openpyxl.load_workbook(CONTROLE, data_only=True)
    out: dict[tuple[int, int], tuple[list[str], Decimal]] = {}
    for row in wb['Alugueis'].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        try:
            cond = int(float(row[0])); unit = int(float(row[1]))
        except (TypeError, ValueError):
            continue
        if cond not in (1, 2):
            continue
        names = [norm(n) for n in str(row[2] or '').split('/') if norm(n)]
        val = Decimal(str(row[3])).quantize(Decimal('0.01')) if row[3] is not None else Decimal('0.00')
        out[(cond, unit)] = (names, val)
    return out


def read_cond3() -> dict[tuple[int, int], tuple[list[str], Decimal]]:
    # nomes do controle (cond 3) + valores do Alugueis unidades.xlsx
    wbc = openpyxl.load_workbook(CONTROLE, data_only=True)
    names_by_unit: dict[int, list[str]] = {}
    for row in wbc['Alugueis'].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        try:
            cond = int(float(row[0])); unit = int(float(row[1]))
        except (TypeError, ValueError):
            continue
        if cond == 3:
            names_by_unit[unit] = [norm(n) for n in str(row[2] or '').split('/') if norm(n)]
    wbv = openpyxl.load_workbook(COND3_VALS, data_only=True)
    out: dict[tuple[int, int], tuple[list[str], Decimal]] = {}
    for row in wbv['Sheet1'].iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            unit = int(float(row[0]))
        except (TypeError, ValueError):
            continue
        val = Decimal(str(row[1])).quantize(Decimal('0.01')) if row[1] is not None else Decimal('0.00')
        out[(3, unit)] = (names_by_unit.get(unit, []), val)
    return out


def gen_charges(db, lease: Lease):
    cur = month_start(lease.start_date)
    while cur <= UP_TO and cur <= month_start(lease.end_date):
        exists = db.scalar(select(RentCharge).where(RentCharge.lease_id == lease.id, RentCharge.reference_month == cur))
        if not exists:
            c = RentCharge(lease_id=lease.id, reference_month=cur,
                           due_date=date(cur.year, cur.month, min(lease.due_day, 28)),
                           amount_due=lease.monthly_rent, amount_paid=Decimal('0.00'),
                           status=ChargeStatus.pending)
            recompute_charge(c); db.add(c)
        cur = add_months(cur, 1)


def run():
    units = {**read_cond12(), **read_cond3()}
    for ku, (names, start) in CONTRATOS_2026.items():  # override inquilino (mantém valor do controle)
        base = units.get(ku, ([], Decimal('0.00')))
        units[ku] = (names, base[1])

    db = SessionLocal()
    try:
        # limpa contratos/cobranças e as unidades by_unit dos condomínios (mantém Inteiro)
        db.execute(delete(RentCharge))
        db.execute(delete(LeaseTenant))
        db.execute(delete(Lease))
        db.commit()
        props = {p.name: p for p in db.scalars(select(Property)).all()}
        for p in props.values():
            if p.rental_mode == RentalMode.by_unit:
                for u in list(p.units):
                    db.delete(u)
        db.commit()

        def tenant_by_name(name: str) -> Tenant | None:
            return db.scalar(select(Tenant).where(func.lower(Tenant.full_name) == name.lower()))

        leases = 0
        # condomínios by_unit
        for (cond, unit), (names, val) in sorted(units.items()):
            prop = props.get(f'Condomínio {cond}')
            if not prop:
                continue
            u = Unit(property_id=prop.id, name=f'Unidade {unit:02d}', base_rent=val)
            db.add(u); db.flush()
            if not names:
                continue
            start = CONTRATOS_2026.get((cond, unit), (None, date(2026, 1, 1)))[1]
            lease = Lease(unit_id=u.id, start_date=start, duration_months=12,
                          end_date=add_months(start, 12) - timedelta(days=1),
                          monthly_rent=val if val > 0 else Decimal('1000.00'), due_day=10)
            recompute_lease(lease); db.add(lease); db.flush()
            for j, nm in enumerate(names):
                t = tenant_by_name(nm)
                if t:
                    db.add(LeaseTenant(lease_id=lease.id, tenant_id=t.id, is_primary=(j == 0)))
            gen_charges(db, lease); leases += 1

        # imóveis inteiros
        for pname, (tenant_name, val) in WHOLE.items():
            prop = props.get(pname)
            if not prop or not prop.units:
                continue
            u = prop.units[0]
            u.base_rent = val
            if not tenant_name:
                continue
            t = tenant_by_name(tenant_name)
            if not t:
                t = Tenant(full_name=tenant_name, notes=f'Inquilino do imóvel inteiro: {pname}')
                db.add(t); db.flush()
            start = date(2026, 1, 1)
            lease = Lease(unit_id=u.id, start_date=start, duration_months=12,
                          end_date=add_months(start, 12) - timedelta(days=1),
                          monthly_rent=val, due_day=10)
            recompute_lease(lease); db.add(lease); db.flush()
            db.add(LeaseTenant(lease_id=lease.id, tenant_id=t.id, is_primary=True))
            gen_charges(db, lease); leases += 1

        recompute_units(db)
        db.commit()
        nleases = db.scalar(select(func.count(Lease.id)))
        ncharges = db.scalar(select(func.count(RentCharge.id)))
        nunits = db.scalar(select(func.count(Unit.id)))
    finally:
        db.close()

    print(f'Unidades: {nunits} | Contratos: {nleases} | Cobranças: {ncharges} (até {UP_TO:%Y-%m})')


if __name__ == '__main__':
    run()
